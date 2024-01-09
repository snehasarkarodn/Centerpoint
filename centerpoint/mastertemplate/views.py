from django.shortcuts import render
from django.http import HttpResponse, HttpResponseNotFound, HttpResponseServerError
import pandas as pd
import os
import uuid
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from .models import ProcessedData, SheetUpdate
import time 
from datetime import datetime
import ast

def process_english(english_string):
    unique_values = []
    seen_values = set()
    for value in english_string.split("|"):
        if value not in seen_values:
            seen_values.add(value)
            unique_values.append(value)
    return unique_values


def color_code_columns(sheet, selected_df, mandatory_fixed):
    header_values = sheet[1]
    mandatory_columns = mandatory_fixed.columns

    for idx, header in enumerate(header_values, start=1):
        col_letter = get_column_letter(idx)

        if header.value in mandatory_columns:
            is_mandatory = mandatory_fixed.at[0, header.value] == 'Yes'
        elif header.value in selected_df["Field Name"].values:
            row_index = selected_df["Field Name"].eq(header.value).idxmax()
            is_mandatory = selected_df.at[row_index, "Mandatory"] == 'yes'
        else:
            continue

        cell = sheet.cell(row=1, column=idx)

        if is_mandatory:
            cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        else:
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow

        if header.value in selected_df["Field Name"].values:
            field_type = selected_df.at[row_index, "Field Type"]
            if field_type == 'select' or field_type == 'multi-select':
                cell.font = openpyxl.styles.Font(color="FF0000")


def add_hidden_data_validation(sheet, col_idx, col_len, hidden_col_idx):
    dv = DataValidation(type="list", formula1=f'hidden_sheet!${get_column_letter(hidden_col_idx)}$2:${get_column_letter(hidden_col_idx)}${col_len + 1}')
    # Apply data validation to the entire column in the second row
    for r in range(2, 3):
        sheet.add_data_validation(dv)
        dv.add(sheet.cell(row=r, column=col_idx))


def add_temp_data_validation(sheet, idx, max_temp_len, hidden_col_idx):
    dv = DataValidation(type="list", formula1=f'hidden_sheet1!${get_column_letter(hidden_col_idx)}$2:${get_column_letter(hidden_col_idx)}${max_temp_len + 1}')
    # Apply data validation to the entire column in the second row
    for r in range(2, 3):
        sheet.add_data_validation(dv)
        dv.add(sheet.cell(row=r, column=idx))

def add_formula(sheet, country, header_suffix, col_letter, input_col):
    country_columns = {
        'UAE': ['Price', 'Color Price', 'SKU Price', 'Concept Delivery'],
        'KSA': ['Price', 'Color Price', 'SKU Price', 'Concept Delivery'],
        'QAT': ['Price', 'Color Price', 'SKU Price', 'Concept Delivery'],
        'KWT': ['Price', 'Color Price', 'SKU Price', 'Concept Delivery']
    }
    col_suffixes = country_columns[country]
    if header_suffix not in col_suffixes:
        for r in range(2, 3):
            sheet[f"{col_letter}{r}"].value = f'=IF(OR({input_col[0]}{r}<>"",{input_col[1]}{r}<>"",{input_col[2]}{r}<>""),"Yes","No")'



def copy_and_modify_master_temp(selected_df, mandatory_fixed, testdata, template_dropdown, selected_values_str, guide_df, merged_temp):
    start =time.time()
    selected_sheet = "CP Content_Template"
    master_temp = pd.read_excel(testdata, sheet_name=selected_sheet)
    selected_sheet = "CP_Content_Template"
    id = str(uuid.uuid4())[:8]
    unique_id = f"ODNCP_{datetime.now().strftime('%d%m%y')}_{id}"
    output_folder = os.path.join("mastertemplate", "stored_data")
    output_path = os.path.join(output_folder, f"{selected_sheet}_{id}.xlsx")
    selected_values = ast.literal_eval(selected_values_str)
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        master_temp.to_excel(writer, sheet_name="Listing", index=False)

    wb = openpyxl.load_workbook(output_path)
    sheet = wb.active

    hidden_sheet = wb.create_sheet("hidden_sheet")
    hidden_sheet.sheet_state = "hidden"

    hidden_sheet1 = wb.create_sheet("hidden_sheet1")
    hidden_sheet1.sheet_state = "hidden"

    guide_sheet =wb.create_sheet("Guidelines")
    guide_data = guide_df.values.tolist()

    for row_index, row_data in enumerate(guide_data):
        for col_index, value in enumerate(row_data):
            guide_sheet.cell(row=row_index + 1, column=col_index + 1, value=value)

    for list_name in selected_values:
        selected_row = merged_temp[merged_temp["Template Name"]== list_name]
        selected_row["English"] = selected_row["English"].astype(str)
        selected_row["Processed_English"] = selected_row["English"].apply(lambda x: process_english(x))
        selected_row.drop(columns=["English","Template Name"],inplace=True)
        transposed_data = selected_row.transpose().reset_index(drop=True)
        data_to_write = transposed_data.values.tolist()
        sheet_add = wb.create_sheet(list_name)
        for row in data_to_write[:-1]:
            sheet_add.append(row)
        for col_values in data_to_write[-1:]:
            for col_index, value in enumerate(col_values):
                for row_index, val in enumerate(value, start=5):
                    sheet_add.cell(row=row_index, column=col_index + 1, value=val)


    processed_english_df = pd.DataFrame()
    max_length = 0

    # add columns as per selected category
    for index, row in selected_df.iterrows():
        field_type_value = row["Field Name"]
        if field_type_value not in sheet[1]:
            sheet.cell(row=1, column=sheet.max_column + 1, value=field_type_value)

        if row.Processed_English != ['nan'] and len(row.Processed_English) > 0:
            max_length = max(max_length, len(row.Processed_English))

    for index, row in selected_df.iterrows():
        field_type_value = row["Field Name"]

        if row.Processed_English != ['nan'] and len(row.Processed_English) > 0:
            padded_values = row.Processed_English + [None] * (max_length - len(row.Processed_English))
            processed_english_df[field_type_value] = padded_values

    # Putting data in the hidden sheet
    headers = list(processed_english_df.columns)
    for c_idx, header in enumerate(headers, start=1):
        hidden_sheet.cell(row=1, column=c_idx, value=header)

    for r_idx, row in enumerate(processed_english_df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            hidden_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Putting data in the hidden sheet1
    headers1 = list(template_dropdown.columns)
    for c_idx, col_name in enumerate(template_dropdown.columns, start=1):
        hidden_sheet1.cell(row=1, column=c_idx, value=col_name)

    for r_idx, row in template_dropdown.iterrows():
        for c_idx, value in enumerate(row, start=1):
            hidden_sheet1.cell(row=r_idx + 2, column=c_idx, value=value)
    wb.save(output_path)

    # Adding Data Validation
    hidden = pd.read_excel(output_path, sheet_name='hidden_sheet')
    temp_hidden = pd.read_excel(output_path, sheet_name='hidden_sheet1')
    sheet = wb.active
    barcode_col_idx = None
    for idx, header in enumerate(sheet[1], start=1):
        if header.value in headers:
            hidden_col_idx = headers.index(header.value) + 1 if header.value in headers else None
            if header.value in hidden.columns:
                col_len = hidden[header.value].count()

            if hidden_col_idx is not None:
                for r in range(2, hidden_sheet.max_row + 1):
                    add_hidden_data_validation(sheet, idx, col_len, hidden_col_idx)
            else:
                for r in range(2, sheet.max_row + 1):
                    sheet.cell(row=r, column=idx, value=None)

        elif header.value in headers1:
            hidden_col_idx = headers1.index(header.value) + 1 if header.value in headers1 else None

            if header.value in temp_hidden.columns:
                max_temp_len = temp_hidden[header.value].count()

            if hidden_col_idx is not None:
                for r in range(2, hidden_sheet1.max_row + 1):
                    add_temp_data_validation(sheet, idx, max_temp_len, hidden_col_idx)
            else:
                for r in range(2, sheet.max_row + 1):
                    sheet.cell(row=r, column=idx, value=None)

        elif header.value == "Base Product ID":
            col_letter = get_column_letter(idx)
            for r in range(2, 3):
                sheet[f"{col_letter}{r}"].value = f'=SUBSTITUTE(D{r}, " ", "")& "CP" & TEXT(TODAY(), "DD-MM-YYYY")'
        
        elif header.value == "Barcode":
            barcode_col_idx = idx

        elif header.value.endswith(('UAE', 'KSA', 'QAT', 'KWT')):
            columns_letter = {
                'UAE': ['J', 'K', 'L'],
                'KSA': ['M', 'N', 'O'],
                'QAT': ['P', 'Q', 'R'],
                'KWT': ['S', 'T', 'U']
            }
            country_code = header.value[-3:]
            header_suffix = header.value[:-3].strip()
            col_letter = get_column_letter(idx)
            if country_code in columns_letter:
                add_formula(sheet, country_code, header_suffix, col_letter, columns_letter[country_code])
    
    if barcode_col_idx is not None:
        for r in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=r, column=barcode_col_idx)
            formula = f'COUNTIF(${get_column_letter(barcode_col_idx)}$2:${get_column_letter(barcode_col_idx)}${sheet.max_row}, ${get_column_letter(barcode_col_idx)}{r})>1'
            rule = openpyxl.formatting.rule.FormulaRule(
                formula=[formula],
                stopIfTrue=True,
                fill=PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            )
            sheet.conditional_formatting.add(f'{get_column_letter(barcode_col_idx)}{r}', rule)

    color_code_columns(sheet, selected_df, mandatory_fixed)

    wb.save(output_path)
    wb = openpyxl.load_workbook(output_path)
    output_path=output_path.replace('.xlsx','.xls')
    wb.save(output_path)

    end =time.time()
    tot_time = end-start

    processed_data_instance = ProcessedData.objects.create(
        unique_id=unique_id,
        output_path=output_path,
        selected_values=selected_values_str,
        filename = f"{selected_sheet}_{id}.xls",
        created_by = "ODN",
        duration_of_creation = (tot_time)
        )

    return output_path


def main_func(request):
    data_records = ProcessedData.objects.all()
    testdata = os.path.join("mastertemplate", "Centerpoint_master_template", "centrepoint_Template and attribute.xlsx")
    merged_temp = pd.read_excel(testdata, sheet_name="Attribute and Values")
    template_dropdown = pd.read_excel(testdata, sheet_name="Template_dropdown_values")
    mandatory_fixed = pd.read_excel(testdata, sheet_name="Template_Mandatory")
    guidelines = pd.read_excel(testdata, sheet_name="Guidelines")

    dropdown_values = [i.strip() for i in merged_temp["Template Name"].unique()]
    selected_df = pd.DataFrame()
    guide_df = pd.DataFrame()

    if request.method == 'POST':
        selected_values_list = request.POST.getlist('selected_items')
        selected_values_str = selected_values_list[0]
        selected_values = ast.literal_eval(selected_values_str)

        for selected_template in selected_values:
            selected_rows = merged_temp[merged_temp["Template Name"]== selected_template]
            guide_row=guidelines[guidelines["Template Name"]== selected_template]
            if not selected_rows.empty:
                selected_df = pd.concat([selected_df, selected_rows], ignore_index=True)
            if not guide_row.empty:
                guide_df = pd.concat([guide_df, guide_row], ignore_index=True)

        selected_df.drop(columns=["Template Name"], inplace=True)
        selected_df["English"] = selected_df["English"].astype(str)
        selected_df["English"] = (selected_df.groupby("Field Name")["English"].transform(lambda x: "|".join(x.unique())))
        selected_df = selected_df.drop_duplicates(subset="Field Name")
        selected_df["Processed_English"] = selected_df["English"].apply(lambda x: process_english(x))

        output_path = copy_and_modify_master_temp(selected_df, mandatory_fixed, testdata, template_dropdown, selected_values_str, guide_df, merged_temp)

        data_records = ProcessedData.objects.all()

        if os.path.exists(output_path):
            return render(request, 'mastertemplate/user_interface1.html', {
                'dropdown_values': dropdown_values,
                'output_path': output_path.replace("\\", "/"),
                'data_records': data_records
            })
        else:
            return HttpResponse("File not found.")

    # Render the initial form page
    return render(request, 'mastertemplate/user_interface1.html', {'dropdown_values': dropdown_values, 'data_records': data_records})


def download_template(request, file_path):
    file_path = file_path.replace("/", "\\")
    file_name = os.path.basename(file_path)
    if not file_path:
        return HttpResponseNotFound("File path is missing")
    if not os.path.isfile(file_path):
        return HttpResponseNotFound("File not found")
    try:
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    except Exception as e:
        return HttpResponseServerError("An error occurred while processing the request")
    

def download_latest_sheet(request):
    file_path = os.path.join("mastertemplate", "Centerpoint_master_template", "centrepoint_Template and attribute.xlsx")
    file_name = os.path.basename(file_path)
    if not file_path:
        return HttpResponseNotFound("File path is missing")
    if not os.path.isfile(file_path):
        return HttpResponseNotFound("File not found")
    try:
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    except Exception as e:
        return HttpResponseServerError("An error occurred while processing the request")

def update_sheet(request):
    data_records = SheetUpdate.objects.all()
    if request.method == 'POST':
        start=time.time()
        sheet_id = request.POST.get('sheet_id', '')
        print(sheet_id)
        ver_name=""

        # Load sheets from Google Sheets
        sheet = pd.ExcelFile(f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx")

        # Load sheets from the local workbook
        testdata = os.path.join("mastertemplate", "Centerpoint_master_template", "centrepoint_Template and attribute.xlsx")
        testdata_workbook = pd.ExcelFile(testdata)
        testdata_sheets = testdata_workbook.sheet_names

        # Read all sheets from the local workbook
        all_local_sheets = {sheet_name: pd.read_excel(testdata_workbook, sheet_name=sheet_name) for sheet_name in testdata_sheets}

        # Check for differences
        differences = []
        sheets_name = []
        output_path =""

        for sheet_name in testdata_sheets:
            if sheet_name in sheet.sheet_names:
                sheet1 = pd.read_excel(sheet, sheet_name=sheet_name)

                if sheet_name in all_local_sheets:
                    sheet2 = all_local_sheets[sheet_name]

                    if not sheet1.equals(sheet2):
                        
                        # Create a backup with timestamp
                        wb = openpyxl.load_workbook(testdata)
                        output_path=testdata.replace('centrepoint_Template and attribute.xlsx',f"CP_Temp&attr_backup_{datetime.now().strftime('%Y%m%d')}.xlsx")
                        ver_name = f"CP_Temp&attr_backup_{datetime.now().strftime('%Y%m%d')}.xlsx"
                        wb.save(output_path)

                        # Update the local sheet if there are differences
                        all_local_sheets[sheet_name] = sheet1
                        differences.append(f"Sheet '{sheet_name}' has been updated.")
                        sheets_name.append(sheet_name)
                    else:
                        differences.append(f"No changes required for sheet '{sheet_name}'.")
                else:
                    differences.append(f"Sheet '{sheet_name}' not found in the local workbook.")
            else:
                differences.append(f"Sheet '{sheet_name}' not found in the Google Sheets workbook.")

        # Add any new sheets from Google Sheets to the local workbook
        new_sheets = set(sheet.sheet_names) - set(all_local_sheets.keys())
        for new_sheet_name in new_sheets:
            all_local_sheets[new_sheet_name] = pd.read_excel(sheet, sheet_name=new_sheet_name)
            wb = openpyxl.load_workbook(testdata)
            output_path=testdata.replace('centrepoint_Template and attribute.xlsx',f"CP_Temp&attr_backup_{datetime.now().strftime('%Y%m%d')}.xlsx")
            ver_name = f"CP_Temp&attr_backup_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(output_path)
            sheets_name.append(new_sheet_name)
            differences.append(f"New sheet '{new_sheet_name}' has been added to the local workbook.")

        # Write all sheets back to the local workbook
        with pd.ExcelWriter(testdata, engine='xlsxwriter') as writer:
            for sheet_name, sheet_data in all_local_sheets.items():
                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

        response_text = "\n".join(differences)
        end=time.time()
        tot_time = end-start

        if ver_name != "":
            updatesheet_data_instance = SheetUpdate.objects.create(
            file_version = ver_name,
            edited_by = "ODN",
            file_path = output_path,
            duration_of_update = (tot_time),
            workbook_id = sheet_id,
            edited_sheets = sheets_name
            )
        else:
            pass
        data_records = SheetUpdate.objects.all()

        if output_path != "":
            return render(request, 'mastertemplate/update_sheet.html', {'response_text': response_text, 'data_records': data_records, 'output_path':output_path})
        else:
            return render(request, 'mastertemplate/update_sheet.html', {'response_text': response_text, 'data_records': data_records})

    return render(request, 'mastertemplate/update_sheet.html', {'data_records': data_records})
