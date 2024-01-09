import os
from django.http import JsonResponse, HttpResponse, HttpResponseNotFound, HttpResponseServerError
from django.shortcuts import render
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.utils import get_column_letter
import numpy as np
from .models import QualityCheckRecord
import time 
from datetime import datetime

def process_english(english_string):
    unique_values = []
    seen_values = set()
    for value in english_string.split("|"):
        if value not in seen_values:
            seen_values.add(value)
            unique_values.append(value)
    return unique_values

def barcode_conditional_format(worksheet, max_rows):
    target_column_name = "Barcode"
    column_number = None
    for column in worksheet.iter_cols(min_row=1, max_row=1):
        for cell in column:
            if cell.value == target_column_name:
                column_number = cell.column
                break
    column_letter = get_column_letter(column_number)
    duplicate_style = NamedStyle(name='duplicate_style')
    duplicate_style.fill = PatternFill(start_color='C72939', end_color='C72939', fill_type='solid')
    worksheet.conditional_formatting.add(f'{column_letter}2:AX{max_rows}', FormulaRule(formula=[f'COUNTIF(${column_letter}$2:${column_letter}${max_rows}, {column_letter}2)>1'], fill=duplicate_style.fill))
                                                  
def process_file(request):
    data_records = QualityCheckRecord.objects.all()
    if request.method == 'POST' and request.FILES.get('excelFile'):
        start=time.time()
        excel_file = request.FILES['excelFile']
        processed_file_name = excel_file.name
        file_name_without_extension = processed_file_name.rsplit('.', 1)[0] 
        id = file_name_without_extension[-8:]
        unique_id = f"ODNCP_{datetime.now().strftime('%d%m%y')}_{id}"
        output_folder = os.path.join("quality_check", "checking_data")
        output_path = os.path.abspath(os.path.join(output_folder, processed_file_name))
        output_path=output_path.replace('.xls','.xlsx')
        with open(output_path, 'wb') as destination:
            for chunk in excel_file.chunks():
                destination.write(chunk)
        
        excel_data = pd.read_excel(output_path, sheet_name='Listing')
        n_rows = excel_data.shape[0]
        excel_data['Remarks'] = ''
        file_path = os.path.join("mastertemplate", "Centerpoint_master_template", "centrepoint_Template and attribute.xlsx")
        temp_file = pd.read_excel(file_path, sheet_name='Attribute and Values')
        temp_mand = pd.read_excel(file_path, sheet_name='Template_Mandatory')
        yes_values_list = [column for column in temp_mand.columns if temp_mand[column].iloc[0] == "Yes"]
        mandatory_col=[]
        missed_header=''
        df_list = [] 
        # Missing header 
        for template_name in excel_data['Template Name'].unique():
            template_columns = temp_file[temp_file['Template Name'] == template_name]['Field Name'].tolist()
            mandatory_columns = temp_file[(temp_file['Template Name'] == template_name) & (temp_file['Mandatory'] == 'yes')]['Field Name']
            df = temp_file[(temp_file['Template Name'] == template_name) & ((temp_file['Field Type'] == 'select') | (temp_file['Field Type'] == 'multi-select'))][["Template Name", "Field Name", "English"]]
            mandatory_col.extend(mandatory_columns)
            missing_columns = [col for col in template_columns if col not in excel_data.columns]
            if missing_columns:
                missing_text = f' {{"{template_name}": {missing_columns}}} |'
                missed_header += missing_text.strip()

            df["English"] = df["English"].astype(str)
            df["Processed_English"] = df["English"].apply(lambda x: process_english(x))
            df_list.append(df)  
        final_df = pd.concat(df_list, ignore_index=True)
        excel_data.loc[0, 'Missing Header'] = missed_header.strip()

        # detecting duplicated bar code
        duplicate_barcode = excel_data[excel_data.duplicated(subset=['Barcode'], keep=False)]
        for _, row in duplicate_barcode.iterrows():
            duplicate_rows = excel_data[excel_data['Barcode'] == row['Barcode']].index
            modified_duplicate_rows = [element + 2 for element in duplicate_rows.tolist()]
            remark_text = f'Duplicate values in rows{modified_duplicate_rows} |'
            excel_data.loc[duplicate_rows, 'Remarks'] = remark_text
        
        # replacing null with blank for mandatory headers for selected category     
        for column_name in mandatory_col:
            excel_data[column_name] = excel_data[column_name].replace(np.nan, '-Blank-')
            empty_cells = excel_data[excel_data[column_name] == '-Blank-'].index
            if len(empty_cells) > 0:
                remark_text = f' Found Empty cell in {column_name} | '
                excel_data.loc[empty_cells, 'Remarks'] += remark_text.lstrip()

        # adding remarks for cell values with null data in mandatory header for fixed headers   
        for column_name in yes_values_list:
            empty_cells = excel_data[excel_data[column_name].isnull()].index
            if len(empty_cells) > 0:
                remark_text = f'Found Empty cell in {column_name} (mandatory header) |'
                excel_data.loc[empty_cells, 'Remarks'] += remark_text.lstrip()

        # Marking Values not in predifined list
        found_cells = [] 
        for _, row in excel_data.iterrows():
            template_name = row['Template Name']
            for field_name in final_df.loc[final_df['Template Name'] == template_name, 'Field Name']:
                if field_name in excel_data.columns:
                    cell_value = row[field_name]
                    if (cell_value is not None) and (str(cell_value).strip() != "") and (not pd.isna(cell_value)):
                        processed_values = final_df.loc[(final_df['Template Name'] == template_name)&(final_df['Field Name'] == field_name), 'Processed_English'].iloc[0]
                        if type(cell_value) == float:
                                cell_value = str(int(cell_value))
                        if cell_value not in processed_values:
                            remark_text = f'{cell_value} of {field_name} is not from the predefined list |'
                            excel_data.at[_, 'Remarks'] += remark_text

                            found_cells.append((_ + 2, excel_data.columns.get_loc(field_name) + 1))
                        else:
                            pass
                    else:
                        remark_text = f'{field_name} is empty |'
                        excel_data.at[_, 'Remarks'] += remark_text

                        found_cells.append((_ + 2, excel_data.columns.get_loc(field_name) + 1))
                            
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            excel_data.to_excel(writer, sheet_name='FinalQC', index=False)

        workbook = load_workbook(output_path)
        worksheet = workbook['FinalQC']
        max_rows = worksheet.max_row
        barcode_conditional_format(worksheet, max_rows)
        
        red_fill = PatternFill(start_color='C72939', end_color='C72939', fill_type='solid')
        for column in worksheet.iter_cols():
            column_header = column[0].value
            if column_header in mandatory_columns: # fill red in blank cell for mandatory headers for selected category  
                for cell in column[1:]:
                    if cell.value == "-Blank-":
                        cell.fill = red_fill

            elif column_header in yes_values_list:  # fill red for cell values with null data in mandatory header for fixed headers 
                for cell in column[1:]:
                    if cell.value is None:
                        cell.fill = red_fill

        fill_color = PatternFill(start_color="C72939", end_color="C72939", fill_type="solid")
        for row, col in found_cells:
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill_color

        workbook.save(output_path)
        wb = openpyxl.load_workbook(output_path)
        output_path=output_path.replace('.xlsx','.xls')
        wb.save(output_path)

        end=time.time()
        tot_time = end-start

        quality_check_data_instance = QualityCheckRecord.objects.create(
        unique_id = unique_id,
        file_path = output_path,
        file_name = processed_file_name,
        num_records = n_rows,
        qc_processing_time = tot_time,
        qc_done_by = "ODN"
        )
        data_records = QualityCheckRecord.objects.all()
        return render(request, 'quality_check/qc_interface.html', {  #'quality_check/qc_interface.html'
                'output_path': output_path.replace("\\", "/"),
                'data_records': data_records
            })
    return render(request, 'quality_check/qc_interface.html', {'data_records': data_records})  # 'quality_check/qc_interface.html'

def download_template(request, file_path):
    file_path = file_path.replace("/", "\\")
    file_name = os.path.basename(file_path)
    if not file_path:
        return HttpResponseNotFound("File path is missing")
    if not os.path.isfile(file_path):
        return HttpResponseNotFound("File not found")
    try:
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
    except Exception as e:
        return HttpResponseServerError("An error occurred while processing the request")